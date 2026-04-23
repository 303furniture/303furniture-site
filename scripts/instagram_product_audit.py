#!/usr/bin/env python3
"""Build a one-sheet Instagram product posting audit workbook."""

from __future__ import annotations

import csv
import html
import re
import time
import unicodedata
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "reports"
OUTPUT_XLSX = REPORT_DIR / "instagram_product_audit.xlsx"
OUTPUT_PRODUCTS_CSV = REPORT_DIR / "instagram_product_audit_products.csv"
OUTPUT_POSTS_CSV = REPORT_DIR / "instagram_posts.csv"

JST = timezone(timedelta(hours=9))
SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9,ja;q=0.8",
    }
)


@dataclass
class Product:
    brand: str
    collection: str
    category: str
    name: str
    url: str
    source: str
    variant_count: int = 1
    notes: list[str] = field(default_factory=list)


@dataclass
class Post:
    code: str
    url: str
    date: str
    timestamp: int
    caption: str
    title: str
    brands: list[str]


BRAND_TERMS = {
    "KETTAL": ["kettal", "@kettal", "#kettal"],
    "ETHIMO": ["ethimo", "@ethimo", "#ethimo"],
    "POINT": ["point1920", "@point1920", "#point1920", "point"],
    "MANUTTI": ["manutti", "@manutti", "#manutti"],
    "HARBOUR": ["harbour", "@harbouroutdoor", "#harbour"],
    "VONDOM": ["vondom", "@vondom", "#vondom"],
}

KETTAL_COLLECTIONS = [
    "Altar", "Anda", "Arc", "Band", "Base Kitchen", "Basket", "Bela lamp",
    "Bitta Lounge", "Grand Bitta", "Bitta", "Boma", "Cala", "Cottage", "Daybed",
    "Dots", "Eolias", "Giro", "Half Dome", "Il Colonnato", "Insula", "Kari",
    "Landscape", "Loden", "Maia", "Mesh", "Meteo", "Mia", "Molo", "Net", "Objects",
    "Pad", "Park Life", "Passage", "Pavilion V", "Pavilion H & L", "Plumon",
    "Ringer", "Riva", "Roll", "Soga rugs", "Superfan", "Tilos", "Tou", "VDL",
    "Village", "Vimini",
]

KETTAL_CATEGORIES = {
    "Pavilions": "https://www.kettal.com/living/en/architecture/pavilions",
    "Architecture Daybeds": "https://www.kettal.com/living/en/architecture/daybeds",
    "Chairs": "https://www.kettal.com/living/en/furnitures/chairs",
    "Barstools": "https://www.kettal.com/living/en/furnitures/barstools",
    "Clubs": "https://www.kettal.com/living/en/furnitures/clubs",
    "Daybeds": "https://www.kettal.com/living/en/furnitures/daybeds",
    "Tables": "https://www.kettal.com/living/en/furnitures/tables",
    "Sofas": "https://www.kettal.com/living/en/furnitures/sofas",
    "Poufs": "https://www.kettal.com/living/en/furnitures/pouffs",
    "Footstools": "https://www.kettal.com/living/en/furnitures/footstool",
    "Sun loungers": "https://www.kettal.com/living/en/furnitures/sunloungers",
    "Parasols": "https://www.kettal.com/living/en/furnitures/parasol",
    "Floor lamps": "https://www.kettal.com/living/en/lighting/floor-lamp",
    "Table lamps": "https://www.kettal.com/living/en/lighting/table-lamp",
    "Lamps": "https://www.kettal.com/living/en/objects/lamp",
    "Planters": "https://www.kettal.com/living/en/objects/planter",
    "Fire pits": "https://www.kettal.com/living/en/objects/fire-pit",
    "Kitchens": "https://www.kettal.com/living/en/objects/kitchen",
    "Rugs": "https://www.kettal.com/living/en/objects/rugs",
    "Object parasols": "https://www.kettal.com/living/en/objects/parasols",
    "Covers": "https://www.kettal.com/living/en/objects/accesories",
    "Cushions": "https://www.kettal.com/living/en/objects/cushions",
    "Fans": "https://www.kettal.com/living/en/objects/fan",
    "Sideboards": "https://www.kettal.com/living/en/objects/sideboards",
}

VONDOM_CATEGORIES = [
    "outdoor-rugs", "lamps", "planters", "saucers", "benches", "counters",
    "lounge-chairs", "decorative-cushions", "daybed", "dining-tables",
    "high-tables", "low-tables", "objects", "canopies", "chairs", "sofas",
    "stools-ottomans", "sun-loungers",
]


def fetch(url: str, *, timeout: int = 45, retries: int = 2) -> requests.Response:
    last_error = None
    for attempt in range(retries + 1):
        try:
            response = SESSION.get(url, timeout=timeout, allow_redirects=True)
            response.raise_for_status()
            return response
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt < retries:
                time.sleep(1.2 + attempt)
    raise RuntimeError(f"failed to fetch {url}: {last_error}")


def clean_text(value: str) -> str:
    value = html.unescape(value or "")
    value = re.sub(r"<[^>]+>", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def title_from_slug(slug: str) -> str:
    slug = re.sub(r"-\d+$", "", slug).replace("_", "-")
    words = [part for part in slug.split("-") if part]
    return " ".join(word.upper() if word.upper() in {"VDL", "UFO"} else word.capitalize() for word in words)


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def norm_contains(text_norm: str, needle: str) -> bool:
    needle_norm = normalize(needle)
    return bool(needle_norm and f" {needle_norm} " in f" {text_norm} ")


def infer_collection(name: str, collections) -> str:
    name_norm = normalize(name)
    matches = []
    for collection in collections:
        collection_norm = normalize(collection)
        if collection_norm and (name_norm.startswith(collection_norm) or f" {collection_norm} " in f" {name_norm} "):
            matches.append(collection)
    return max(matches, key=lambda item: len(normalize(item))) if matches else ""


def parse_xml_locs(xml_text: str) -> list[str]:
    root = ET.fromstring(xml_text)
    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    return [loc.text or "" for loc in root.findall(".//sm:loc", ns)]


def parse_meta_content(html_text: str, attr_name: str) -> str:
    patterns = [
        rf'<meta\s+property=["\']{re.escape(attr_name)}["\']\s+content=["\']([^"\']+)["\']',
        rf'<meta\s+content=["\']([^"\']+)["\']\s+property=["\']{re.escape(attr_name)}["\']',
        rf'<meta\s+name=["\']{re.escape(attr_name)}["\']\s+content=["\']([^"\']+)["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html_text, flags=re.I)
        if match:
            return clean_text(match.group(1))
    return ""


def parse_title(html_text: str) -> str:
    match = re.search(r"<title[^>]*>(.*?)</title>", html_text, flags=re.I | re.S)
    return clean_text(match.group(1)) if match else ""


def extract_instagram_title(caption: str) -> str:
    match = re.search(r"【([^】]+)】", caption or "")
    return clean_text(match.group(1)) if match else ""


def detect_brands(caption: str, title: str) -> list[str]:
    text = normalize(f"{title} {caption}")
    title_norm = normalize(title)
    brands = []
    for brand, terms in BRAND_TERMS.items():
        if brand == "POINT":
            if title_norm.startswith("point") or "point1920" in text or "@point1920" in caption.lower():
                brands.append(brand)
            continue
        for term in terms:
            if normalize(term) in text:
                brands.append(brand)
                break
    return brands


def fetch_instagram_posts() -> list[Post]:
    endpoint = "https://www.instagram.com/api/v1/feed/user/in_out.furniture/username/"
    headers = {"x-ig-app-id": "936619743392459", "Referer": "https://www.instagram.com/in_out.furniture/"}
    posts: list[Post] = []
    next_max_id = None
    seen_codes: set[str] = set()
    for _ in range(12):
        params = {"count": "12"}
        if next_max_id:
            params["max_id"] = next_max_id
        response = SESSION.get(endpoint, params=params, headers=headers, timeout=45)
        response.raise_for_status()
        payload = response.json()
        for item in payload.get("items", []):
            code = item.get("code")
            if not code or code in seen_codes:
                continue
            caption = (item.get("caption") or {}).get("text") or ""
            timestamp = int(item.get("taken_at") or item.get("device_timestamp") or 0)
            date = datetime.fromtimestamp(timestamp, timezone.utc).astimezone(JST).strftime("%Y-%m-%d")
            title = extract_instagram_title(caption)
            posts.append(Post(code, f"https://www.instagram.com/p/{code}/", date, timestamp, caption, title, detect_brands(caption, title)))
            seen_codes.add(code)
        if not payload.get("more_available"):
            break
        next_max_id = payload.get("next_max_id")
        if not next_max_id:
            break
        time.sleep(0.7)
    return sorted(posts, key=lambda post: post.timestamp)


def fetch_titles(urls: list[str], parser, workers: int = 8) -> list[Product]:
    products: list[Product] = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(parser, url): url for url in urls}
        for future in as_completed(futures):
            try:
                product = future.result()
            except Exception as exc:  # noqa: BLE001
                url = futures[future]
                products.append(Product("", "", "", title_from_slug(urlparse(url).path.rstrip("/").split("/")[-1]), url, "fetch failed", notes=[str(exc)[:120]]))
                continue
            if product:
                products.append(product)
    return products


def match_slug_collection(slug: str, collection_slugs: list[str], collection_names: list[str]) -> str:
    best_index = None
    best_len = 0
    slug_norm = slug.replace("-", " ")
    for index, collection_slug in enumerate(collection_slugs):
        collection_norm = collection_slug.replace("-", " ")
        if slug_norm.startswith(collection_norm) and len(collection_slug) > best_len:
            best_index = index
            best_len = len(collection_slug)
    return collection_names[best_index] if best_index is not None else title_from_slug(slug.split("-")[0])


def match_slug_anywhere(slug: str, collection_slugs: list[str], collection_names: list[str]) -> str:
    best_index = None
    best_len = 0
    slug_norm = slug.replace("-", " ")
    for index, collection_slug in enumerate(collection_slugs):
        collection_norm = collection_slug.replace("-", " ")
        if f" {collection_norm} " in f" {slug_norm} " and len(collection_slug) > best_len:
            best_index = index
            best_len = len(collection_slug)
    return collection_names[best_index] if best_index is not None else ""


def scrape_ethimo() -> list[Product]:
    sitemap = fetch("https://www.ethimo.com/sitemaps/sitemap_products.xml").text
    urls = sorted({url for url in parse_xml_locs(sitemap) if "/en/product/" in url})

    def parse_product(url: str) -> Product:
        page = fetch(url, timeout=30, retries=1)
        title = parse_title(page.text).replace(" | Ethimo", "").strip() or title_from_slug(urlparse(page.url).path.rstrip("/").split("/")[-1])
        collection = title.split(" - ")[-1].strip() if " - " in title else ""
        return Product("ETHIMO", collection, "", title, page.url, "official sitemap")

    products = fetch_titles(urls, parse_product, workers=8)
    for product in products:
        if not product.brand:
            product.brand = "ETHIMO"
            product.source = "official sitemap"
    return dedupe_products(products)


def scrape_point() -> list[Product]:
    sitemap = fetch("https://www.point1920.com/sitemap.xml").text
    locs = parse_xml_locs(sitemap)
    product_urls = sorted({url for url in locs if "point1920.com/products/" in url})
    collection_slugs = [urlparse(url).path.rstrip("/").split("/")[-1] for url in locs if "point1920.com/coleccion/" in url]
    collection_names = [title_from_slug(slug) for slug in collection_slugs]

    def parse_product(url: str) -> Product:
        page = fetch(url, timeout=35, retries=1)
        final_slug = urlparse(page.url).path.rstrip("/").split("/")[-1]
        raw_title = parse_meta_content(page.text, "og:title") or parse_title(page.text)
        category = raw_title.replace("| POINT, SL", "").strip().title()
        collection = match_slug_collection(final_slug, collection_slugs, collection_names)
        name = category if collection and normalize(category).startswith(normalize(collection)) else f"{collection} {category}".strip()
        return Product("POINT", collection, category, name, page.url, "official sitemap")

    products = fetch_titles(product_urls, parse_product, workers=8)
    for product in products:
        if not product.brand:
            product.brand = "POINT"
            product.source = "official sitemap"
    return dedupe_products(products)


def scrape_manutti() -> list[Product]:
    sitemap = fetch("https://www.manutti.com/sitemap.xml").text
    urls = parse_xml_locs(sitemap)
    collection_urls = [url for url in urls if "manutti.com/en/collections/" in url]
    collection_slugs = [urlparse(url).path.rstrip("/").split("/")[-1] for url in collection_urls]
    collection_names = [title_from_slug(slug) for slug in collection_slugs]
    product_urls = []
    for url in urls:
        if "manutti.com/en/products/" not in url:
            continue
        parts = urlparse(url).path.strip("/").split("/")
        if len(parts) == 5:
            product_urls.append(url)
    products = []
    for url in sorted(set(product_urls)):
        parts = urlparse(url).path.strip("/").split("/")
        category = title_from_slug(parts[3])
        slug = parts[-1]
        collection = match_slug_anywhere(slug, collection_slugs, collection_names)
        type_slug = slug
        if collection:
            collection_slug = normalize(collection).replace(" ", "-")
            index = slug.find(collection_slug)
            if index > 0:
                type_slug = slug[:index].rstrip("-")
        type_slug = re.sub(r"^outdoor-", "", type_slug)
        type_name = title_from_slug(type_slug)
        name = f"{collection} {type_name}".strip() if collection else title_from_slug(slug)
        products.append(Product("MANUTTI", collection, category, name, url, "official sitemap, normalized from URL slug"))
    return dedupe_products(products)


def scrape_harbour() -> list[Product]:
    sitemap_index = fetch("https://shopharbour.com/sitemap.xml").text
    sitemaps = parse_xml_locs(sitemap_index)
    product_sitemap = next(url for url in sitemaps if "sitemap_products_1.xml" in url and "/el/" not in url)
    collection_sitemap = next(url for url in sitemaps if "sitemap_collections_1.xml" in url and "/el/" not in url)
    collection_names = [title_from_slug(urlparse(url).path.rstrip("/").split("/")[-1].replace("-collection", "")) for url in parse_xml_locs(fetch(collection_sitemap).text) if "/collections/" in url]
    root = ET.fromstring(fetch(product_sitemap).text)
    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9", "image": "http://www.google.com/schemas/sitemap-image/1.1"}
    products = []
    for url_node in root.findall("sm:url", ns):
        loc_node = url_node.find("sm:loc", ns)
        if loc_node is None or not loc_node.text or "/products/" not in loc_node.text:
            continue
        title_node = url_node.find("image:image/image:title", ns)
        name = clean_text(title_node.text if title_node is not None else "") or title_from_slug(urlparse(loc_node.text).path.rstrip("/").split("/")[-1])
        collection = infer_collection(name, collection_names)
        products.append(Product("HARBOUR", collection, "", name, loc_node.text, "official Shopify sitemap"))
    return dedupe_products(products)


def scrape_vondom() -> list[Product]:
    collection_sitemap = fetch("https://www.vondom.com/collection-sitemap.xml").text
    collection_names = [
        title_from_slug(urlparse(url).path.rstrip("/").split("/")[-1].replace("-collection", ""))
        for url in parse_xml_locs(collection_sitemap)
        if "/collections/" in url and "/es/" not in url and url.rstrip("/") != "https://www.vondom.com/collections"
    ]
    products = []
    for category_slug in VONDOM_CATEGORIES:
        category = title_from_slug(category_slug)
        page = fetch(f"https://www.vondom.com/products-category/{category_slug}", timeout=50, retries=1).text
        for href, inner in re.findall(r'<a[^>]+href=["\']([^"\']*www\.vondom\.com/products/[^"\']+)["\'][^>]*>(.*?)</a>', page, flags=re.I | re.S):
            name = clean_text(inner)
            if not name or name.lower() in {"products", "all"}:
                continue
            name = clean_text(re.sub(r"\bRef:\s*\d+\b", "", re.sub(r"\bView set\b", "", name, flags=re.I), flags=re.I))
            collection = infer_collection(name, collection_names)
            products.append(Product("VONDOM", collection, category, name, href, "official product category page"))
    return dedupe_products(products)


def scrape_kettal() -> list[Product]:
    products = []
    forbidden_slugs = {
        "architecture", "furnitures", "lighting", "objects", "finishes", "collections", "designers", "projects",
        "our-history", "flag-stores", "authorized-dealers", "client-support", "general-guarantee", "our-materials",
        "contact", "news", "privacy-policy",
    }
    for category, url in KETTAL_CATEGORIES.items():
        reader_url = "https://r.jina.ai/http://r.jina.ai/http://" + url
        markdown = fetch(reader_url, timeout=70, retries=1).text
        for name, href in re.findall(r"\[([^\]]+)\]\((https://www\.kettal\.com/living/en/[^)]+)\)", markdown):
            parsed = urlparse(href)
            parts = parsed.path.strip("/").split("/")
            if len(parts) != 3 or parts[-1] in forbidden_slugs or len(clean_text(name)) < 3:
                continue
            collection = infer_collection(clean_text(name), KETTAL_COLLECTIONS)
            products.append(Product("KETTAL", collection, category, clean_text(name), href, "official category page via Jina Reader"))
    return dedupe_products(products)


def dedupe_products(products: list[Product]) -> list[Product]:
    grouped: dict[tuple[str, str, str, str], Product] = {}
    for product in products:
        key = (product.brand, normalize(product.collection), normalize(product.category), normalize(product.name))
        if key in grouped:
            grouped[key].variant_count += product.variant_count
        else:
            grouped[key] = product
    result = list(grouped.values())
    for product in result:
        if product.variant_count > 1:
            product.notes.append(f"{product.variant_count} variant URLs consolidated")
    return sorted(result, key=lambda item: (item.brand, item.collection, item.category, item.name))


def audit_products(products: list[Product], posts: list[Post]) -> list[dict]:
    posts_by_brand: dict[str, list[Post]] = defaultdict(list)
    for post in posts:
        for brand in post.brands:
            posts_by_brand[brand].append(post)
    rows = []
    for product in products:
        exact_posts = []
        collection_posts = []
        for post in posts_by_brand.get(product.brand, []):
            post_norm = normalize(f"{post.title} {post.caption}")
            if product.name and norm_contains(post_norm, product.name):
                exact_posts.append(post)
            elif product.collection and norm_contains(post_norm, product.collection):
                collection_posts.append(post)
        matched_posts = exact_posts or collection_posts
        if exact_posts:
            status, reason = "投稿済", "商品名一致"
        elif collection_posts:
            status, reason = "コレクション投稿あり", "コレクション名一致"
        else:
            status, reason = "未投稿", ""
        dates = [post.date for post in matched_posts]
        rows.append(
            {
                "投稿状況": status,
                "ブランド": product.brand,
                "コレクション": product.collection,
                "カテゴリ": product.category,
                "商品名": product.name,
                "商品URL": product.url,
                "Instagram投稿URL": "\n".join([post.url for post in matched_posts[:5]]),
                "初回投稿日": min(dates) if dates else "",
                "投稿数": len(matched_posts),
                "最終投稿日": max(dates) if dates else "",
                "判定根拠": reason,
                "備考": "; ".join(product.notes),
                "_source": product.source,
            }
        )
    return sorted(rows, key=lambda row: (str(row["ブランド"]), str(row["コレクション"]), str(row["カテゴリ"]), str(row["商品名"])))


def write_csv(rows: list[dict], posts: list[Post]) -> None:
    product_headers = ["投稿状況", "ブランド", "コレクション", "カテゴリ", "商品名", "商品URL", "Instagram投稿URL", "初回投稿日", "投稿数", "最終投稿日", "判定根拠", "備考", "_source"]
    with OUTPUT_PRODUCTS_CSV.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=product_headers)
        writer.writeheader()
        writer.writerows(rows)
    with OUTPUT_POSTS_CSV.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=["投稿日", "投稿URL", "タイトル", "ブランド", "キャプション"])
        writer.writeheader()
        for post in posts:
            writer.writerow({"投稿日": post.date, "投稿URL": post.url, "タイトル": post.title, "ブランド": ", ".join(post.brands), "キャプション": post.caption})


def write_workbook(rows: list[dict], posts: list[Post]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "投稿管理"
    title_fill = PatternFill("solid", fgColor="1F2933")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    status_fills = {
        "投稿済": PatternFill("solid", fgColor="C6E0B4"),
        "コレクション投稿あり": PatternFill("solid", fgColor="FFE699"),
        "未投稿": PatternFill("solid", fgColor="E7E6E6"),
        "判定保留": PatternFill("solid", fgColor="F4CCCC"),
    }
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet["A1"] = "Instagram投稿 商品管理"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    sheet["A1"].fill = title_fill
    sheet.merge_cells("A1:L1")
    sheet["A2"] = "対象Instagram"
    sheet["B2"] = "https://www.instagram.com/in_out.furniture/"
    sheet["A3"] = "作成日時"
    sheet["B3"] = datetime.now(JST).strftime("%Y-%m-%d %H:%M")
    sheet["A4"] = "取得投稿数"
    sheet["B4"] = len(posts)
    sheet["A5"] = "判定メモ"
    sheet["B5"] = "商品名一致は投稿済、コレクション名のみ一致はコレクション投稿ありとして判定"
    sheet.merge_cells("B5:L5")
    summary_start = 7
    summary_headers = ["ブランド", "総商品数", "投稿済", "コレクション投稿あり", "未投稿"]
    for col, header in enumerate(summary_headers, 1):
        cell = sheet.cell(summary_start, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
    counts_by_brand = defaultdict(Counter)
    for row in rows:
        counts_by_brand[str(row["ブランド"])][str(row["投稿状況"])] += 1
        counts_by_brand[str(row["ブランド"])] ["total"] += 1
    for offset, brand in enumerate(sorted(counts_by_brand), 1):
        data = counts_by_brand[brand]
        values = [brand, data["total"], data["投稿済"], data["コレクション投稿あり"], data["未投稿"]]
        for col, value in enumerate(values, 1):
            cell = sheet.cell(summary_start + offset, col, value)
            cell.border = border
    table_start = summary_start + len(counts_by_brand) + 3
    headers = ["投稿状況", "ブランド", "コレクション", "カテゴリ", "商品名", "商品URL", "Instagram投稿URL", "初回投稿日", "投稿数", "最終投稿日", "判定根拠", "備考"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(table_start, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(rows, table_start + 1):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row_index, col, row[header])
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header in {"商品URL", "Instagram投稿URL"} and row[header]:
                cell.hyperlink = str(row[header]).split("\n")[0]
                cell.style = "Hyperlink"
        sheet.cell(row_index, 1).fill = status_fills.get(str(row["投稿状況"]), PatternFill())
    widths = {"A": 18, "B": 13, "C": 20, "D": 24, "E": 38, "F": 58, "G": 48, "H": 13, "I": 10, "J": 13, "K": 18, "L": 34}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = sheet.cell(table_start + 1, 1)
    sheet.auto_filter.ref = f"A{table_start}:L{table_start + len(rows)}"
    sheet.row_dimensions[1].height = 24
    for row_number in range(table_start + 1, table_start + len(rows) + 1):
        sheet.row_dimensions[row_number].height = 42
    workbook.save(OUTPUT_XLSX)


def main() -> None:
    REPORT_DIR.mkdir(exist_ok=True)
    print("Fetching Instagram posts...")
    posts = fetch_instagram_posts()
    print(f"  posts: {len(posts)}")
    scrapers = [("KETTAL", scrape_kettal), ("ETHIMO", scrape_ethimo), ("POINT", scrape_point), ("MANUTTI", scrape_manutti), ("HARBOUR", scrape_harbour), ("VONDOM", scrape_vondom)]
    all_products: list[Product] = []
    for brand, scraper in scrapers:
        print(f"Fetching {brand} products...")
        products = scraper()
        print(f"  products: {len(products)}")
        all_products.extend(products)
    rows = audit_products(all_products, posts)
    write_csv(rows, posts)
    write_workbook(rows, posts)
    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_PRODUCTS_CSV}")
    print(f"Wrote {OUTPUT_POSTS_CSV}")


if __name__ == "__main__":
    main()
